import openpyxl
from openpyxl.utils import get_column_letter
import requests
from PIL import Image
from io import BytesIO  #바이너리 파일로 사용하기 위한 라이브러리
from openpyxl.drawing.image import Image    #엑셀에 이미지 삽입하는 라이브러리

def set_withs(ws):
    column_withs={} #각 열의 최대 문자열 길이를 저장할 딕셔너리

    for row in ws.iter_rows():
        for cell in row :
            cell_lengh = len(cell.value) # 셀 값의 문자열 길이 계산
            column_letter = get_column_letter(cell.column)

            #기존 최대값과 비교하여 큰 값을 저장
            if column_letter in column_withs : 
                column_withs[column_letter] = max(column_withs[column_letter],cell_lengh)

            else :
                column_withs[column_letter]=cell_lengh

    # 각 열의 너비를 설정
    for column_letter,width in column_withs.items():
        ws.column_dimensions[column_letter].width = width+3

#열너비 조정
wb = openpyxl.load_workbook("3_danawa_crawling_result.xlsx")
ws = wb.active
# set_withs(ws)

#이미지 삽입
for row in ws.iter_rows(min_row=2,min_col=4):
    for cell in row :
        # print(cell.value)
        # break
        res = requests.get(cell.value)
        image_data = BytesIO(res.content) #이미지를 엑셀에 넣을 수 있도록 변환
        image_data.width = 100
        image_data.height = 200
        ws.column_dimensions["E"].width=30
        ws.row_dimensions[cell.row].height=200
        img = Image(image_data)
        ws.add_image(img,'e'+str(cell.row))

wb.save("3_danawa_crawling_result.xlsx")